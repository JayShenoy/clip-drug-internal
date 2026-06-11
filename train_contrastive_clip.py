"""
train_contrastive_clip.py
─────────────────────────
Dual-Tower CLIP for aligning small-molecule SMILES with cellular proteome traces.
Dataset:  Mitchell et al. Nat Biotechnol 2023  ("A proteome-wide atlas of drug MOA")

Architecture
────────────
Chemical Tower  : frozen ChemBERTa-77M-MTR  →  2-layer MLP head  →  D=128
Proteome Tower  : Trace Encoder — each compound's log2FC values (z-scored, NaN→0)
                  mapped as scalar node features onto the protein-protein
                  interactome (MOESM6, Pearson correlations across 875 compounds,
                  35 936 edges, 2 387 proteins).

                  Three interchangeable encoder backends, selected via
                  cfg.encoder_type:

                  'deepset'     — element-wise MLP per protein (+ optional
                                  learnable protein-identity embedding) → global
                                  mean pool → 2-layer projection MLP → D=128
                                  Fully permutation-invariant; no graph prior.

                  'gnn'         — 3-layer multi-head GATConv over the PPI graph
                                  → global mean pool → 2-layer projection MLP
                                  → D=128.  Learns to weight edges dynamically.

                  'transformer' — 2-layer multi-head self-attention (no positional
                                  encoding; order-invariant) → mean pool →
                                  2-layer projection MLP → D=128.  Captures
                                  non-local protein interactions without a fixed
                                  graph prior.

Loss            : symmetric InfoNCE (CLIP-style) with learnable temperature
Optimiser       : AdamW + CosineAnnealingLR
Evaluation      : Top-1 / Top-5 retrieval accuracy on validation set
"""

# ── stdlib ───────────────────────────────────────────────────────────────────
import dataclasses
import math
import random
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

# ── third-party ──────────────────────────────────────────────────────────────
import numpy as np
import openpyxl
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy import stats
from torch.utils.data import DataLoader, Dataset, Subset
from transformers import AutoModel, AutoTokenizer

from torch_geometric.data import Data as PyGData, Batch as PyGBatch
from torch_geometric.nn import GATConv, global_mean_pool


# ────────────────────────────────────────────────────────────────────────────
# 0.  Config
# ────────────────────────────────────────────────────────────────────────────

@dataclasses.dataclass
class TrainConfig:
    # ── encoder selection ────────────────────────────────────────────────────
    # One of: 'mlp' | 'deepset' | 'gnn' | 'transformer'
    encoder_type: str = "mlp"

    # ── shared dims ──────────────────────────────────────────────────────────
    latent_dim: int  = 128    # shared CLIP embedding dimension D
    mlp_hidden: int  = 512    # hidden width of projection MLPs

    # ── training ─────────────────────────────────────────────────────────────
    seed:          int   = 42
    batch_size:    int   = 64
    epochs:        int   = 50
    lr:            float = 3e-4
    weight_decay:  float = 1e-2
    val_frac:      float = 0.20
    checkpoint:    str   = "best_proteome_clip.pt"

    # ── chemical tower ───────────────────────────────────────────────────────
    chemberta_model: str = "DeepChem/ChemBERTa-77M-MTR"
    max_smiles_len:  int = 128

    # ── PPI graph ────────────────────────────────────────────────────────────
    ppi_file: str = "data/41587_2022_1539_MOESM6_ESM.xlsx"

    # ── mlp options ──────────────────────────────────────────────────────────
    mlp_enc_hidden: int = 1024    # hidden width of the flat proteome MLP

    # ── deepset options ──────────────────────────────────────────────────────
    deepset_hidden:  int  = 256   # width of per-protein element MLP
    use_prot_emb:    bool = True  # learnable protein-identity embedding

    # ── gnn options ──────────────────────────────────────────────────────────
    gnn_hidden: int = 256
    gnn_heads:  int = 4

    # ── transformer options ──────────────────────────────────────────────────
    tf_hidden:  int = 256   # per-token hidden dim inside transformer
    tf_heads:   int = 4
    tf_layers:  int = 2
    tf_dropout: float = 0.1


cfg = TrainConfig()

# ── device ───────────────────────────────────────────────────────────────────
if torch.backends.mps.is_available():
    DEVICE = torch.device("mps")
elif torch.cuda.is_available():
    DEVICE = torch.device("cuda")
else:
    DEVICE = torch.device("cpu")
print(f"Using device: {DEVICE}")
print(f"Encoder type : {cfg.encoder_type}")


# ────────────────────────────────────────────────────────────────────────────
# 1.  Reproducibility
# ────────────────────────────────────────────────────────────────────────────

def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


set_seed(cfg.seed)


# ────────────────────────────────────────────────────────────────────────────
# 2.  Data loading
# ────────────────────────────────────────────────────────────────────────────

def load_fc_matrix() -> pd.DataFrame:
    """Load log2FC matrix (proteins × compounds). 0 → NaN."""
    print("Loading fold-change matrix …")
    df = pd.read_csv(
        "data/41587_2022_1539_MOESM3_ESM.csv",
        encoding="latin-1",
        index_col=0,
    )
    df = df.drop(columns=["UniprotID"], errors="ignore")
    df = df.apply(pd.to_numeric, errors="coerce")
    df = df.replace(0, np.nan)
    df.columns = [c.replace("\xa0", " ").strip() for c in df.columns]
    # Collapse duplicate gene-name rows (multiple isoforms) by nanmean
    n_before = len(df)
    df = df.groupby(df.index).mean()
    n_after = len(df)
    if n_before != n_after:
        print(f"  Collapsed {n_before - n_after} duplicate gene rows → {n_after} unique proteins")
    print(f"  FC matrix: {df.shape[0]} proteins × {df.shape[1]} compounds")
    return df


def load_smiles_table() -> pd.DataFrame:
    """Load supplementary compound table (MOESM4) that contains SMILES."""
    print("Loading SMILES table …")
    wb = openpyxl.load_workbook("data/41587_2022_1539_MOESM4_ESM.xlsx", read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df["fc_key"] = (
        df["Compound Name"].str.strip() + "_" + df["Primary Target"].str.strip()
    )
    df = df.dropna(subset=["SMILES"])
    print(f"  SMILES table: {len(df)} compounds with valid SMILES")
    return df


PEARSON_CACHE = Path("data/pearson_cache.npz")


def build_pearson_graph(
    fc_df: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, list[str]]:
    """
    Compute compound-compound Pearson correlations (cached).
    Kept for cluster_compounds.py compatibility; not used by the GNN encoder.
    """
    print("Building Pearson correlation graph …")
    compounds = fc_df.columns.tolist()
    mat_f = fc_df.values.astype(np.float32)

    if PEARSON_CACHE.exists():
        print(f"  Loading cached Pearson matrices from {PEARSON_CACHE} …")
        cache = np.load(PEARSON_CACHE, allow_pickle=True)
        if cache["compounds"].tolist() == compounds:
            print("  Done (from cache).")
            return cache["corr_matrix"], cache["count_matrix"], mat_f, compounds
        print("  Cache compound list mismatch — recomputing …")

    n = len(compounds)
    corr_matrix  = np.full((n, n), np.nan, dtype=np.float32)
    count_matrix = np.zeros((n, n), dtype=np.int32)

    for i in range(n):
        xi = mat_f[:, i]
        for j in range(i, n):
            xj = mat_f[:, j]
            mask = ~(np.isnan(xi) | np.isnan(xj))
            cnt = int(mask.sum())
            count_matrix[i, j] = count_matrix[j, i] = cnt
            if cnt >= 3:
                r, _ = stats.pearsonr(xi[mask], xj[mask])
                corr_matrix[i, j] = corr_matrix[j, i] = float(r)
        if i % 100 == 0:
            print(f"  Pearson {i}/{n} …")

    print("  Done.")
    np.savez_compressed(
        PEARSON_CACHE,
        corr_matrix=corr_matrix,
        count_matrix=count_matrix,
        compounds=np.array(compounds),
    )
    print(f"  Pearson matrices cached → {PEARSON_CACHE}")
    return corr_matrix, count_matrix, mat_f, compounds


def load_ppi_graph(fc_df: pd.DataFrame) -> tuple[PyGData, list[str]]:
    """
    Load the protein-protein Pearson correlation network (MOESM6), intersect
    with the FC matrix proteins, and return a PyGData with placeholder node
    features and the canonical protein ordering.
    """
    print("Loading protein-protein interactome (MOESM6) …")
    wb = openpyxl.load_workbook(cfg.ppi_file, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    fc_proteins = set(fc_df.index.tolist())
    src_list, dst_list, w_list = [], [], []
    prot_set: set[str] = set()
    for row in rows[1:]:
        p1, p2, r = str(row[0]), str(row[1]), float(row[2])
        if p1 in fc_proteins and p2 in fc_proteins:
            prot_set.add(p1); prot_set.add(p2)
            src_list.append(p1); dst_list.append(p2); w_list.append(r)

    prot_order = sorted(prot_set)
    prot_idx   = {p: i for i, p in enumerate(prot_order)}
    n = len(prot_order)

    srcs = [prot_idx[p] for p in src_list]
    dsts = [prot_idx[p] for p in dst_list]
    edge_index = torch.tensor([srcs + dsts, dsts + srcs], dtype=torch.long)
    edge_attr  = torch.tensor(w_list + w_list, dtype=torch.float32).unsqueeze(1)
    x          = torch.zeros(n, 1, dtype=torch.float32)  # placeholder

    ppi_graph = PyGData(x=x, edge_index=edge_index, edge_attr=edge_attr)
    ppi_graph.num_nodes = n
    print(
        f"  PPI graph: {n} proteins (FC∩PPI overlap), "
        f"{len(w_list)} undirected edges, continuous Pearson r weights"
    )
    return ppi_graph, prot_order


# ────────────────────────────────────────────────────────────────────────────
# 3.  Dataset  —  uniform output regardless of encoder type
# ────────────────────────────────────────────────────────────────────────────

class ProteomeCLIPDataset(Dataset):
    """
    Each item is a (smiles_str, fc_tensor) pair where fc_tensor has shape
    (N_prot, 1) — the z-scored log2FC value for each protein in prot_order,
    with NaN → 0.  This shape is consumed directly by all three encoder types.
    """

    def __init__(
        self,
        smiles_list:   list[str],
        compound_keys: list[str],   # FC matrix column names
        fc_df:         pd.DataFrame,
        prot_order:    list[str],
    ):
        self.smiles = smiles_list

        # Align FC matrix rows to canonical protein order; fill missing → NaN
        fc_sub = fc_df.reindex(index=prot_order)[compound_keys]  # (N_prot, B)
        fc_np  = fc_sub.values.astype(np.float32)

        # Z-score each protein row across the compounds in this dataset
        mu    = np.nanmean(fc_np, axis=1, keepdims=True)
        sigma = np.nanstd( fc_np, axis=1, keepdims=True) + 1e-8
        fc_z  = np.nan_to_num((fc_np - mu) / sigma, nan=0.0)  # (N_prot, B)

        # Store as (B, N_prot, 1)
        self.fc_tensors = torch.from_numpy(
            fc_z.T[:, :, np.newaxis].astype(np.float32)
        )

    def __len__(self) -> int:
        return len(self.smiles)

    def __getitem__(self, idx: int):
        return self.smiles[idx], self.fc_tensors[idx]   # (N_prot, 1)


# ────────────────────────────────────────────────────────────────────────────
# 4.  Collate  —  encoder-aware but Dataset-agnostic
# ────────────────────────────────────────────────────────────────────────────

def make_collate(ppi_graph: PyGData, encoder_type: str):
    """
    Returns a collate_fn.

    For 'gnn': stamps each compound's (N_prot, 1) FC tensor as node features
    onto the fixed PPI graph topology and returns a PyGBatch.

    For 'deepset' / 'transformer': stacks FC tensors into (B, N_prot, 1) and
    returns a plain torch.Tensor — no graph object needed.

    In both cases the loader yields:  (smiles_list, proteome_data)
    where proteome_data type depends on the encoder.
    """
    ei = ppi_graph.edge_index
    ea = ppi_graph.edge_attr

    if encoder_type == "gnn":
        def collate(batch):
            smiles = [item[0] for item in batch]
            graphs = [PyGData(x=item[1], edge_index=ei, edge_attr=ea) for item in batch]
            return smiles, PyGBatch.from_data_list(graphs)
    else:
        def collate(batch):
            smiles   = [item[0] for item in batch]
            fc_stack = torch.stack([item[1] for item in batch])  # (B, N_prot, 1)
            return smiles, fc_stack

    return collate


# ────────────────────────────────────────────────────────────────────────────
# 5.  Chemical Tower  (frozen — never modified)
# ────────────────────────────────────────────────────────────────────────────

class ChemicalTower(nn.Module):
    def __init__(self):
        super().__init__()
        print(f"  Loading ChemBERTa from '{cfg.chemberta_model}' …")
        self.tokenizer = AutoTokenizer.from_pretrained(cfg.chemberta_model)
        self.encoder   = AutoModel.from_pretrained(cfg.chemberta_model)
        for param in self.encoder.parameters():
            param.requires_grad = False
        bert_hidden = self.encoder.config.hidden_size  # 384
        self.projector = nn.Sequential(
            nn.Linear(bert_hidden, cfg.mlp_hidden),
            nn.GELU(),
            nn.LayerNorm(cfg.mlp_hidden),
            nn.Linear(cfg.mlp_hidden, cfg.latent_dim),
        )

    def tokenize(self, smiles_batch: list[str]) -> dict:
        return self.tokenizer(
            smiles_batch,
            padding=True,
            truncation=True,
            max_length=cfg.max_smiles_len,
            return_tensors="pt",
        )

    def forward(self, input_ids: torch.Tensor, attention_mask: torch.Tensor) -> torch.Tensor:
        with torch.no_grad():
            out = self.encoder(input_ids=input_ids, attention_mask=attention_mask)
        cls = out.last_hidden_state[:, 0, :]
        return self.projector(cls)   # (B, D)


# ────────────────────────────────────────────────────────────────────────────
# 6.  Proteome Encoder backends
# ────────────────────────────────────────────────────────────────────────────

def _projection_mlp(in_dim: int) -> nn.Sequential:
    """Shared 2-layer projection head used by all three backends."""
    return nn.Sequential(
        nn.Linear(in_dim, cfg.mlp_hidden),
        nn.GELU(),
        nn.LayerNorm(cfg.mlp_hidden),
        nn.Linear(cfg.mlp_hidden, cfg.latent_dim),
    )


# ── 6a. Flat MLP encoder ─────────────────────────────────────────────────────

class MLPEncoder(nn.Module):
    """
    Standard 3-layer MLP over the flattened proteome FC vector.

    Treats the full (N_prot,) z-scored FC profile as a fixed-length input
    vector and maps it directly to D=128 through three fully-connected layers.

    Input : fc  (B, N_prot, 1)
    Output: z   (B, D)
    """

    def __init__(self, n_proteins: int):
        super().__init__()
        H = cfg.mlp_enc_hidden
        self.net = nn.Sequential(
            nn.Linear(n_proteins, H),
            nn.GELU(),
            nn.LayerNorm(H),
            nn.Linear(H, H),
            nn.GELU(),
            nn.LayerNorm(H),
            nn.Linear(H, cfg.latent_dim),
        )

    def forward(self, fc: torch.Tensor) -> torch.Tensor:
        return self.net(fc.squeeze(-1))   # (B, N_prot, 1) → (B, N_prot) → (B, D)


# ── 6b. DeepSet encoder ──────────────────────────────────────────────────────

class DeepSetEncoder(nn.Module):
    """
    Treats the proteome as an unordered set of (protein, FC-value) entries.

    Each protein node is represented by its scalar FC value optionally
    augmented with a learnable protein-identity embedding.  A shared
    element-wise 2-layer MLP maps each node to a hidden vector; global mean
    pooling collapses the set; a final projection MLP outputs D=128.

    Permutation-invariant by construction — no graph topology required.

    Input : fc  (B, N_prot, 1)
    Output: z   (B, D)
    """

    def __init__(self, n_proteins: int):
        super().__init__()
        in_dim = 1
        if cfg.use_prot_emb:
            self.prot_emb = nn.Embedding(n_proteins, cfg.deepset_hidden)
            in_dim = 1 + cfg.deepset_hidden
        else:
            self.prot_emb = None

        # Shared element-wise MLP  (applied identically to every protein node)
        self.phi = nn.Sequential(
            nn.Linear(in_dim, cfg.deepset_hidden),
            nn.GELU(),
            nn.LayerNorm(cfg.deepset_hidden),
            nn.Linear(cfg.deepset_hidden, cfg.deepset_hidden),
            nn.GELU(),
            nn.LayerNorm(cfg.deepset_hidden),
        )
        self.projector = _projection_mlp(cfg.deepset_hidden)

    def forward(self, fc: torch.Tensor) -> torch.Tensor:
        # fc : (B, N_prot, 1)
        B, N, _ = fc.shape
        if self.prot_emb is not None:
            ids   = torch.arange(N, device=fc.device)          # (N,)
            emb   = self.prot_emb(ids).unsqueeze(0).expand(B, -1, -1)  # (B, N, E)
            x_in  = torch.cat([fc, emb], dim=-1)               # (B, N, 1+E)
        else:
            x_in = fc                                           # (B, N, 1)

        h      = self.phi(x_in)                                # (B, N, H)
        pooled = h.mean(dim=1)                                  # (B, H)  — global mean pool
        return self.projector(pooled)                           # (B, D)


# ── 6b. GNN encoder ──────────────────────────────────────────────────────────

class GNNEncoder(nn.Module):
    """
    3-layer multi-head GATConv over the protein-protein interactome.

    Node features are the scalar log2FC value at each protein (1-dim).
    Edge weights are continuous Pearson r (no hard threshold); the GAT learns
    to weight edges dynamically during backpropagation.  Global mean pool over
    all protein nodes → projection MLP → D=128.

    Input : PyGBatch (x shape (B*N_prot, 1), batch vector)
    Output: z (B, D)
    """

    def __init__(self):
        super().__init__()
        H, heads = cfg.gnn_hidden, cfg.gnn_heads
        self.conv1 = GATConv(1,        H, heads=heads, edge_dim=1, concat=True,  dropout=0.1)
        self.conv2 = GATConv(H*heads,  H, heads=heads, edge_dim=1, concat=True,  dropout=0.1)
        self.conv3 = GATConv(H*heads,  H, heads=1,     edge_dim=1, concat=False, dropout=0.1)
        self.norm1 = nn.LayerNorm(H * heads)
        self.norm2 = nn.LayerNorm(H * heads)
        self.norm3 = nn.LayerNorm(H)
        self.projector = _projection_mlp(H)

    def forward(self, x, edge_index, edge_attr, batch) -> torch.Tensor:
        h = F.gelu(self.norm1(self.conv1(x, edge_index, edge_attr=edge_attr)))
        h = F.gelu(self.norm2(self.conv2(h, edge_index, edge_attr=edge_attr)))
        h = F.gelu(self.norm3(self.conv3(h, edge_index, edge_attr=edge_attr)))
        return self.projector(global_mean_pool(h, batch))       # (B, D)


# ── 6c. Transformer (Set Attention) encoder ──────────────────────────────────

class TransformerEncoder(nn.Module):
    """
    Lightweight set transformer over protein tokens.

    No positional encoding — proteins are treated as an unordered set, making
    the encoder permutation-invariant.  Multi-head self-attention allows every
    protein to attend to every other protein, capturing non-local interactions
    without a fixed graph prior.  Mean pool over token dim → projection MLP
    → D=128.

    Input : fc  (B, N_prot, 1)
    Output: z   (B, D)
    """

    def __init__(self):
        super().__init__()
        H = cfg.tf_hidden

        # Project scalar FC → token dimension
        self.input_proj = nn.Linear(1, H)

        # Stack of TransformerEncoder layers (each: MHSA + FFN + LayerNorm)
        encoder_layer = nn.TransformerEncoderLayer(
            d_model=H,
            nhead=cfg.tf_heads,
            dim_feedforward=H * 4,
            dropout=cfg.tf_dropout,
            activation="gelu",
            batch_first=True,   # expects (B, N, H)
            norm_first=True,    # pre-LN for training stability
        )
        self.transformer = nn.TransformerEncoder(
            encoder_layer,
            num_layers=cfg.tf_layers,
        )
        self.projector = _projection_mlp(H)

    def forward(self, fc: torch.Tensor) -> torch.Tensor:
        # fc : (B, N_prot, 1)
        tokens = self.input_proj(fc)            # (B, N_prot, H)
        tokens = self.transformer(tokens)       # (B, N_prot, H)
        pooled = tokens.mean(dim=1)             # (B, H)  — permutation-invariant mean pool
        return self.projector(pooled)           # (B, D)


# ────────────────────────────────────────────────────────────────────────────
# 7.  Factory
# ────────────────────────────────────────────────────────────────────────────

class ProteomeEncoderFactory:
    """
    Instantiates the correct proteome encoder backend from cfg.encoder_type.

    Usage:
        encoder = ProteomeEncoderFactory.build(n_proteins=2387)
    """

    VALID = ("mlp", "deepset", "gnn", "transformer")

    @staticmethod
    def build(n_proteins: int) -> nn.Module:
        t = cfg.encoder_type
        if t not in ProteomeEncoderFactory.VALID:
            raise ValueError(
                f"Unknown encoder_type '{t}'. Choose from {ProteomeEncoderFactory.VALID}."
            )
        if t == "mlp":
            return MLPEncoder(n_proteins=n_proteins)
        if t == "deepset":
            return DeepSetEncoder(n_proteins=n_proteins)
        if t == "gnn":
            return GNNEncoder()
        return TransformerEncoder()


# ────────────────────────────────────────────────────────────────────────────
# 8.  CLIP model
# ────────────────────────────────────────────────────────────────────────────

class ProteomeCLIP(nn.Module):
    def __init__(self, n_proteins: int):
        super().__init__()
        self.chem_tower  = ChemicalTower()
        self.prot_tower  = ProteomeEncoderFactory.build(n_proteins)
        self.encoder_type = cfg.encoder_type
        # Learnable log-temperature: initialised to log(1/0.07) ≈ 2.659
        self.log_temp = nn.Parameter(torch.tensor(math.log(1.0 / 0.07)))

    @property
    def temperature(self) -> torch.Tensor:
        return torch.exp(torch.clamp(self.log_temp, max=math.log(100.0)))

    def encode_proteome(self, proteome_data) -> torch.Tensor:
        """
        Dispatch to the correct encoder signature based on encoder type.

        proteome_data is a PyGBatch for 'gnn', or a (B, N_prot, 1) tensor
        for 'deepset' / 'transformer'.
        """
        if self.encoder_type == "gnn":
            return self.prot_tower(
                proteome_data.x,
                proteome_data.edge_index,
                proteome_data.edge_attr,
                proteome_data.batch,
            )
        else:
            return self.prot_tower(proteome_data)   # (B, N_prot, 1)

    def forward(
        self,
        input_ids:      torch.Tensor,
        attention_mask: torch.Tensor,
        proteome_data,                   # PyGBatch or (B, N_prot, 1) tensor
    ) -> tuple[torch.Tensor, torch.Tensor]:
        z_chem = self.chem_tower(input_ids, attention_mask)  # (B, D)
        z_prot = self.encode_proteome(proteome_data)          # (B, D)
        return F.normalize(z_chem, dim=-1), F.normalize(z_prot, dim=-1)


# ────────────────────────────────────────────────────────────────────────────
# 9.  Loss
# ────────────────────────────────────────────────────────────────────────────

def info_nce_loss(
    z_chem:      torch.Tensor,
    z_prot:      torch.Tensor,
    temperature: torch.Tensor,
) -> torch.Tensor:
    """Symmetric InfoNCE / CLIP loss."""
    B      = z_chem.shape[0]
    logits = (z_chem @ z_prot.T) * temperature
    labels = torch.arange(B, device=z_chem.device)
    return (F.cross_entropy(logits, labels) + F.cross_entropy(logits.T, labels)) / 2.0


# ────────────────────────────────────────────────────────────────────────────
# 10.  Training utilities
# ────────────────────────────────────────────────────────────────────────────

def _forward_batch(model: ProteomeCLIP, smiles_batch, proteome_data):
    """Tokenise SMILES, move everything to DEVICE, run one forward pass."""
    enc        = model.chem_tower.tokenize(list(smiles_batch))
    input_ids  = enc["input_ids"].to(DEVICE)
    attn_mask  = enc["attention_mask"].to(DEVICE)
    if isinstance(proteome_data, PyGBatch):
        proteome_data = proteome_data.to(DEVICE)
    else:
        proteome_data = proteome_data.to(DEVICE)
    return model(input_ids, attn_mask, proteome_data)


def train_one_epoch(
    model:     ProteomeCLIP,
    loader:    DataLoader,
    optimiser: torch.optim.Optimizer,
) -> float:
    model.train()
    total_loss, n_samples = 0.0, 0
    for smiles_batch, proteome_data in loader:
        z_c, z_p = _forward_batch(model, smiles_batch, proteome_data)
        loss = info_nce_loss(z_c, z_p, model.temperature)
        optimiser.zero_grad()
        loss.backward()
        nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
        optimiser.step()
        b = len(smiles_batch)
        total_loss += loss.item() * b
        n_samples  += b
    return total_loss / n_samples


@torch.no_grad()
def eval_one_epoch(model: ProteomeCLIP, loader: DataLoader) -> float:
    model.eval()
    total_loss, n_samples = 0.0, 0
    for smiles_batch, proteome_data in loader:
        z_c, z_p = _forward_batch(model, smiles_batch, proteome_data)
        loss = info_nce_loss(z_c, z_p, model.temperature)
        b = len(smiles_batch)
        total_loss += loss.item() * b
        n_samples  += b
    return total_loss / n_samples


@torch.no_grad()
def retrieval_accuracy(
    model:  ProteomeCLIP,
    loader: DataLoader,
    ks:     tuple[int, ...] = (1, 5),
) -> dict[str, float]:
    """Chemistry-to-proteome Top-K retrieval accuracy on the validation set."""
    model.eval()
    all_z_chem, all_z_prot = [], []
    for smiles_batch, proteome_data in loader:
        z_c, z_p = _forward_batch(model, smiles_batch, proteome_data)
        all_z_chem.append(z_c.cpu())
        all_z_prot.append(z_p.cpu())
    zc  = torch.cat(all_z_chem)
    zp  = torch.cat(all_z_prot)
    sim = zc @ zp.T
    labels = torch.arange(len(zc))
    return {
        f"top{k}": (sim.topk(k, dim=1).indices == labels.unsqueeze(1))
                   .any(1).float().mean().item()
        for k in ks
    }


# ────────────────────────────────────────────────────────────────────────────
# 11.  Main
# ────────────────────────────────────────────────────────────────────────────

def main() -> None:
    set_seed(cfg.seed)

    # ── 11a. Load raw data ───────────────────────────────────────────────────
    fc_df     = load_fc_matrix()
    smiles_df = load_smiles_table()

    # ── 11b. Load protein-protein interactome ────────────────────────────────
    ppi_graph, prot_order = load_ppi_graph(fc_df)
    n_proteins = len(prot_order)

    # ── 11c. Match SMILES ↔ FC compounds ────────────────────────────────────
    fc_col_set = set(fc_df.columns.tolist())
    valid_smiles, valid_fc_keys = [], []
    for _, row in smiles_df.iterrows():
        key = row["fc_key"]
        if key in fc_col_set:
            smi = str(row["SMILES"]).strip()
            if smi and smi.lower() != "nan":
                valid_smiles.append(smi)
                valid_fc_keys.append(key)
    print(f"\nMatched {len(valid_smiles)} SMILES ↔ FC profiles")

    # ── 11d. Dataset + 80/20 split ───────────────────────────────────────────
    dataset = ProteomeCLIPDataset(valid_smiles, valid_fc_keys, fc_df, prot_order)
    n_total  = len(dataset)
    n_val    = max(1, int(n_total * cfg.val_frac))
    indices  = list(range(n_total))
    random.shuffle(indices)
    train_set = Subset(dataset, indices[: n_total - n_val])
    val_set   = Subset(dataset, indices[n_total - n_val :])
    print(f"  Train: {len(train_set)}  |  Val: {len(val_set)}")

    collate_fn   = make_collate(ppi_graph, cfg.encoder_type)
    train_loader = DataLoader(
        train_set, batch_size=cfg.batch_size, shuffle=True,
        num_workers=0, drop_last=True, collate_fn=collate_fn,
    )
    val_loader = DataLoader(
        val_set, batch_size=cfg.batch_size, shuffle=False,
        num_workers=0, drop_last=False, collate_fn=collate_fn,
    )

    # ── 11e. Model ───────────────────────────────────────────────────────────
    print("\nBuilding model …")
    model      = ProteomeCLIP(n_proteins=n_proteins).to(DEVICE)
    n_trainable = sum(p.numel() for p in model.parameters() if p.requires_grad)
    n_frozen    = sum(p.numel() for p in model.chem_tower.encoder.parameters())
    print(f"  Encoder type     : {cfg.encoder_type}")
    print(f"  Trainable params : {n_trainable:,}")
    print(f"  Frozen ChemBERTa : {n_frozen:,}")

    # ── 11f. Optimiser + scheduler ───────────────────────────────────────────
    optimiser = torch.optim.AdamW(
        [p for p in model.parameters() if p.requires_grad],
        lr=cfg.lr, weight_decay=cfg.weight_decay,
    )
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimiser, T_max=cfg.epochs, eta_min=cfg.lr * 0.01,
    )

    # ── 11g. Training loop ───────────────────────────────────────────────────
    best_val_loss = float("inf")
    history       = []
    print(f"\nTraining for {cfg.epochs} epochs …\n")
    header = (
        f"{'Epoch':>5}  {'Train Loss':>10}  {'Val Loss':>10}"
        f"  {'Top-1':>6}  {'Top-5':>6}  {'Temp':>6}"
    )
    print(header)
    print("─" * len(header))

    for epoch in range(1, cfg.epochs + 1):
        train_loss = train_one_epoch(model, train_loader, optimiser)
        val_loss   = eval_one_epoch(model, val_loader)
        retrieval  = retrieval_accuracy(model, val_loader, ks=(1, 5))
        scheduler.step()

        temp_val = model.temperature.item()
        history.append(dict(
            epoch=epoch, train_loss=train_loss, val_loss=val_loss,
            top1=retrieval["top1"], top5=retrieval["top5"],
            temperature=temp_val,
        ))
        print(
            f"{epoch:>5}  {train_loss:>10.4f}  {val_loss:>10.4f}"
            f"  {retrieval['top1']:>6.3f}  {retrieval['top5']:>6.3f}"
            f"  {temp_val:>6.3f}"
        )

        if val_loss < best_val_loss:
            best_val_loss = val_loss
            torch.save(
                {
                    "epoch": epoch,
                    "encoder_type": cfg.encoder_type,
                    "model_state_dict": model.state_dict(),
                    "optimiser_state_dict": optimiser.state_dict(),
                    "val_loss": val_loss,
                    "top1": retrieval["top1"],
                    "top5": retrieval["top5"],
                    "temperature": temp_val,
                    "prot_order": prot_order,
                    "valid_smiles": valid_smiles,
                    "valid_fc_keys": valid_fc_keys,
                },
                cfg.checkpoint,
            )
            print(f"  ✓ New best checkpoint saved → {cfg.checkpoint}")

    # ── 11h. Final summary ───────────────────────────────────────────────────
    best = min(history, key=lambda r: r["val_loss"])
    print(f"\n{'─' * 60}")
    print(f"Encoder     : {cfg.encoder_type}")
    print(f"Best epoch  : {best['epoch']}")
    print(f"Val loss    : {best['val_loss']:.4f}")
    print(f"Top-1 acc   : {best['top1']:.3f}")
    print(f"Top-5 acc   : {best['top5']:.3f}")
    print(f"Checkpoint  : {cfg.checkpoint}")
    pd.DataFrame(history).to_csv("training_history.csv", index=False)
    print("Training history saved → training_history.csv")


if __name__ == "__main__":
    main()
